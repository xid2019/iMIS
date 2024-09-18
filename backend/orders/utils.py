import shutil
import openpyxl
import os
from openpyxl.styles import Border, Side, Alignment

def delete_old_files(directory):
    # Delete all .xlsx files in the given directory that start with 'PO_'.
    for filename in os.listdir(directory):
        if filename.startswith('PO_') and filename.endswith('.xlsx'):
            file_path = os.path.join(directory, filename)
            os.remove(file_path)
            print(f"Deleted: {file_path}")

def create_po_excel(data, shipping_address):
    
    source_path = "./static/PO TemplateLT.xlsx"
    destination_directory = "../results/PO"
    # Delete old files before creating new ones
    delete_old_files(destination_directory)
    
    # split the data by factories
    factory_map_to_data = {}
    for order_line in data:
        factory = order_line[13]
        if factory not in factory_map_to_data:
            factory_map_to_data[factory] = []
        factory_map_to_data[factory].append(order_line)

    for factory in factory_map_to_data:
        data = factory_map_to_data[factory]    
        # Step 1: Copy the Excel file to a new directory
        destination_path = os.path.join(destination_directory, "PO_" + factory+ ".xlsx")
        # Ensure the destination directory exists
        os.makedirs(destination_directory, exist_ok=True)

        # Copy the file
        shutil.copy(source_path, destination_path)

        # Step 2: Open the copied file and fill in the data
        # Load the workbook
        wb = openpyxl.load_workbook(destination_path)
        sheet = wb['Sheet1']
        
        # Define the starting row (where you want to start inserting new rows)
        start_row = 17
        start_col = 2
        col_map = {
            5:2,
            8:3,
            6:5,
            7:10,
            9:12,
            11:13,
            12:14,
            15:6,
            16:7,
            17:17,
            18:11,
            19:15
        }
        total = 0
        # Iterate over the 2D data array
        for row_index, row_data in enumerate(data):
            # Insert a new row at the correct position
            row_to_insert = start_row + row_index
            sheet.insert_rows(row_to_insert)
            sheet.row_dimensions[row_to_insert].height = 20
            # Write data into the newly inserted row
            for col_index, value in enumerate(row_data):
                if col_index in col_map:
                    cell = sheet.cell(row=row_to_insert, column=col_map[col_index])
                    cell.value = value
                    cell.alignment = Alignment(horizontal='center', vertical='center')
            cell = sheet.cell(row=row_to_insert, column=19)
            cell.value = float(sheet.cell(row=row_to_insert, column=3).value) * float(sheet.cell(row=row_to_insert, column=17).value)  
            total += cell.value
            cell.alignment = Alignment(horizontal='center', vertical='center')   

        # add order number
        cell = sheet.cell(row=2, column=18)
        cell.value = data[0][2]

        # add total
        cell = sheet.cell(row=start_row + len(data) + 2, column=18)
        cell.value = total

        # add order date
        cell = sheet.cell(row=start_row + len(data) + 8, column=18)
        cell.value = data[0][3]

        # add air shipping address
        for i in range(9, 13):
            cell = sheet.cell(row=i, column=16)
            cell.value = shipping_address[i-9]
        
        # add supplier info
        sheet.cell(row=9, column=2).value = data[0][21]
        sheet.cell(row=10, column=2).value = data[0][22]
        sheet.cell(row=11, column=2).value = data[0][23]
        sheet.cell(row=12, column=2).value = data[0][24]
        # Define the border style
        border_style = Side(border_style="thin", color="000000")

        # Create a border object with the defined style
        border = Border(left=border_style, right=border_style, top=border_style, bottom=border_style)
        left_border = Border(left=Side(border_style="medium"), right=border_style, top=border_style, bottom=border_style)
        right_border = Border(left=border_style, right=Side(border_style="medium"), top=border_style, bottom=border_style)
        bottom_border = Border(left=border_style, right=border_style, top=border_style, bottom=Side(border_style="medium"))
        
        # add borders
        end_row = start_row + len(data)-1
        end_col = start_col + 17
        for row in range(start_row, end_row):
            for col in range(start_col, end_col):
                cell = sheet.cell(row=row, column=col)
                cell.border = border
        
        # bottom border
        for col in range(start_col, end_col):
            cell = sheet.cell(row=end_row, column=col)
            cell.border = bottom_border

        # Left border
        for row in range(start_row, end_row):
            cell = sheet.cell(row=row, column=start_col)
            cell.border = left_border

        # Right border
        for row in range(start_row, end_row):
            cell = sheet.cell(row=row, column=end_col)
            cell.border = right_border
        
        left_bottom_cell = sheet.cell(row=end_row, column=start_col)
        left_bottom_cell.border = bottom_border = Border(left=Side(border_style="medium"), right=border_style, top=border_style, bottom=Side(border_style="medium"))
        right_bottom_cell = sheet.cell(row=end_row, column=end_col)
        right_bottom_cell.border = bottom_border = Border(left=border_style, right=Side(border_style="medium"), top=border_style, bottom=Side(border_style="medium"))
        # Save the workbook
        wb.save(destination_path)