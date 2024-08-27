from rest_framework.decorators import api_view
from rest_framework.response import Response
from rest_framework import status
from .serializers import OrderSerializer
from order_lines.serializers import OrderLineSerializer
from order_lines.models import OrderLine
from django.db import transaction, IntegrityError
from django.db import connection
from django.db.utils import OperationalError
from orders.models import Order
import shutil
import openpyxl
import os
from openpyxl.styles import Border, Side, Alignment


@api_view(['GET'])
def get_orders(request):
    # Extract query parameters
    customer_id = request.GET.get('customer_id')
    customer_po = request.GET.get('customer_po')
    part_number = request.GET.get('part_number')
    order_date_after = request.GET.get('order_date_after')
    order_date_before = request.GET.get('order_date_before')
    required_date_after = request.GET.get('required_date_after')
    required_date_before = request.GET.get('required_date_after_date_before')
    status = request.GET.get('status')

    # Base SQL query
    sql_query = """
        SELECT
            o.id AS order_id,
            o.customer_id,
            o.customer_po,
            o.order_date,
            ol.id AS orderline_id,
            ol.line_number,
            ol.part_number,
            ol.description,
            ol.quantity,
            ol.ship_via,
            ol.balance,
            ol.required_date,
            ol.confirmed_date,
            ol.factory,
            ol.status
        FROM
            orders_order AS o
        JOIN
            order_lines_orderline AS ol
        ON
            o.id = ol.order_id
    """

    # Initialize an empty list to hold WHERE conditions
    where_clauses = []
    query_params = []

    # Add WHERE clauses based on available query parameters
    if customer_id:
        where_clauses.append("o.customer_id = %s")
        query_params.append(customer_id)

    if customer_po:
        where_clauses.append("o.customer_po = %s")
        query_params.append(customer_po)
    
    if part_number:
        where_clauses.append("ol.part_number = %s")
        query_params.append(part_number)

    if order_date_after:
        where_clauses.append("o.order_date >= %s")
        query_params.append(order_date_after)

    if order_date_before:
        where_clauses.append("o.order_date <= %s")
        query_params.append(order_date_after)

    if required_date_after:
        where_clauses.append("ol.required_date >= %s")
        query_params.append(required_date_after)

    if required_date_before:
        where_clauses.append("ol.required_date <= %s")
        query_params.append(required_date_after)

    if status:
        where_clauses.append("ol.status = %s")
        query_params.append(status)

    # Combine the WHERE clauses
    if where_clauses:
        sql_query += " WHERE " + " AND ".join(where_clauses)

    sql_query += " ORDER BY o.order_date DESC"

    try:
        # Execute the query with parameters
        with connection.cursor() as cursor:
            cursor.execute(sql_query, query_params)
            rows = cursor.fetchall()

        keys = [
            "order_id",
            "customer_id",
            "customer_po",
            "order_date",
            "orderline_id",
            "line_number",
            "part_number",
            "description",
            "quantity",
            "ship_via",
            "balance",
            "required_date",
            "confirmed_date",
            "factory",
            "status"
        ]

        # Convert the result to a list of dictionaries
        result = [dict(zip(keys, row)) for row in rows]

        return Response(result)

    except OperationalError as e:
        return Response({"error": str(e)}, status=500)


@api_view(['GET'])
def get_order(request, order_id):
    # Execute raw SQL query
    with connection.cursor() as cursor:
        cursor.execute("""
            SELECT
                o.id AS order_id,
                o.customer_id,
                o.customer_po,
                o.order_date,
                ol.id AS orderline_id,
                ol.line_number,
                ol.part_number,
                ol.description,
                ol.quantity,
                ol.ship_via,
                ol.required_date,
                ol.status
            FROM
                orders_order AS o
            LEFT JOIN
                order_lines_orderline AS ol
            ON
                o.id = ol.order_id
            WHERE
                o.id = %s
        """, [order_id])
        rows = cursor.fetchall()
    # Process the rows into a format suitable for serialization
    orders = {}
    for row in rows:
        order_id = row[0]
        if order_id not in orders:
            orders[order_id] = {
                'id': order_id,
                'customer_id': row[1],
                'customer_po': row[2],
                'order_date': row[3],
                'order_lines': []
            }
        if row[4]:  # Check if orderline_id is not null
            orders[order_id]['order_lines'].append({
                'id': row[4],
                'line_number': row[5],
                'part_number': row[6],
                'description': row[7],
                'quantity': row[8],
                'ship_via': row[9],
                'required_date': row[10],
                'status': row[11]
            })

    # Convert the dictionary into a list of orders
    order_list = list(orders.values())

    return Response(order_list)


@api_view(['POST'])
def create_order(request):
    with transaction.atomic():
        try:
            customer_po = request.data[0]['customer_po']
        except KeyError:
            return Response({"detail": "Missing Customer PO"}, status=status.HTTP_400_BAD_REQUEST)
        
        try:
            order = Order.objects.get(customer_po=customer_po)
        except Order.DoesNotExist:
            order_data = {
                'customer_id': request.data[0]['customer_id'],
                'customer_po': customer_po,
                'order_date': request.data[0]['order_date'],
                'buyer': request.data[0]['buyer']
            }
            order_serializer = OrderSerializer(data=order_data)
            if order_serializer.is_valid():
                order = order_serializer.save()
            else:
                return Response(order_serializer.errors, status=status.HTTP_400_BAD_REQUEST)
            
        order_lines = []
        for order_line in request.data:
            order_line['order'] = order.id  # Set the order foreign key
            order_line_serializer = OrderLineSerializer(data=order_line)
            if order_line_serializer.is_valid():
                order_line_instance = OrderLine(**order_line_serializer.validated_data)
                order_lines.append(order_line_instance)
            else:
                transaction.set_rollback(True)
                return Response(order_line_serializer.errors, status=status.HTTP_400_BAD_REQUEST)
        
        try:
            OrderLine.objects.bulk_create(order_lines)
        except Exception as e:
            transaction.set_rollback(True)
            return Response({"detail": str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

    # fetch the order with all its order lines
    with connection.cursor() as cursor:
        cursor.execute("""
            SELECT
                o.id AS order_id,
                o.customer_id,
                o.customer_po,
                o.order_date,
                ol.id AS orderline_id,
                ol.line_number,
                ol.part_number,
                ol.description,
                ol.quantity,
                ol.ship_via,
                ol.balance,
                ol.required_date,
                ol.confirmed_date,
                ol.factory,
                ol.status
            FROM
                orders_order AS o
            JOIN
                order_lines_orderline AS ol
            ON
                o.id = ol.order_id
            WHERE
                o.customer_po = %s
        """, [customer_po])
        rows = cursor.fetchall()
        create_po(rows)
    return Response(status=status.HTTP_201_CREATED)

def create_po(data):
    # Step 1: Copy the Excel file to a new directory
    source_path = "./static/PO TemplateLT.xlsx"
    destination_directory = "../results"
    destination_path = os.path.join(destination_directory, "PO.xlsx")

    # Ensure the destination directory exists
    os.makedirs(destination_directory, exist_ok=True)

    # Copy the file
    shutil.copy(source_path, destination_path)

    # Step 2: Open the copied file and fill in the data
    # Load the workbook
    wb = openpyxl.load_workbook(destination_path)
    sheet = wb['Sheet1']
    
    # add order number
    cell = sheet.cell(row=2, column=18)
    cell.value = data[0][2]

    # Define the border style
    border_style = Side(border_style="thin", color="000000")

    # Create a border object with the defined style
    border = Border(left=border_style, right=border_style, top=border_style, bottom=border_style)
    left_border = Border(left=Side(border_style="medium"), right=border_style, top=border_style, bottom=border_style)
    right_border = Border(left=border_style, right=Side(border_style="medium"), top=border_style, bottom=border_style)
    bottom_border = Border(left=border_style, right=border_style, top=border_style, bottom=Side(border_style="medium"))
    
    # Define the starting row (where you want to start inserting new rows)
    start_row = 17
    start_col = 2
    col_map = {
        5:2,
        8:3,
        6:5,
        7:10,
        9:12,
        11:13,
        12:14
    }
    # Iterate over the 2D data array
    for row_index, row_data in enumerate(data):
        # Insert a new row at the correct position
        row_to_insert = start_row + row_index
        sheet.insert_rows(row_to_insert)
        sheet.row_dimensions[row_index].height = 20
        # Write data into the newly inserted row
        for col_index, value in enumerate(row_data):
            if col_index in col_map:
                cell = sheet.cell(row=row_to_insert, column=col_map[col_index])
                cell.value = value
                cell.alignment = Alignment(horizontal='center', vertical='center')
                

    # add borders
    end_row = start_row + len(data)-1
    end_col = start_col + 17
    for row in range(start_row, end_row):
        for col in range(start_col, end_col):
            cell = sheet.cell(row=row, column=col)
            cell.border = border
    
    # bottom border
    for col in range(start_col, end_col):
        cell = sheet.cell(row=end_row, column=col)
        cell.border = bottom_border

    # Left border
    for row in range(start_row, end_row):
        cell = sheet.cell(row=row, column=start_col)
        cell.border = left_border

    # Right border
    for row in range(start_row, end_row):
        cell = sheet.cell(row=row, column=end_col)
        cell.border = right_border
    
    left_bottom_cell = sheet.cell(row=end_row, column=start_col)
    left_bottom_cell.border = bottom_border = Border(left=Side(border_style="medium"), right=border_style, top=border_style, bottom=Side(border_style="medium"))
    right_bottom_cell = sheet.cell(row=end_row, column=end_col)
    right_bottom_cell.border = bottom_border = Border(left=border_style, right=Side(border_style="medium"), top=border_style, bottom=Side(border_style="medium"))
    # Save the workbook
    wb.save(destination_path)

   