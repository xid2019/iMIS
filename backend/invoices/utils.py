import shutil
import openpyxl
import os
from openpyxl.styles import Border, Side, Alignment
from datetime import datetime


def delete_old_files(directory):
    # Delete all .xlsx files in the given directory that start with 'PO_'.
    for filename in os.listdir(directory):
        file_path = os.path.join(directory, filename)
        os.remove(file_path)
        print(f"Deleted: {file_path}")


def create_invoice_excel(data):
    source_path = "./static/Invoice Template.xlsx"
    destination_directory = "../results/Invoice"
    # Delete old files before creating new ones
    delete_old_files(destination_directory)

    os.makedirs(destination_directory, exist_ok=True)

    destination_path = os.path.join(destination_directory, "Invoice.xlsx")

    # Copy the file
    shutil.copy(source_path, destination_path)

    # Load the workbook
    wb = openpyxl.load_workbook(destination_path)
    sheet = wb["Sheet1"]

    order_lines = data["orderLineData"]
    extra_charges = data["extraChargeData"]
    address_data = data["addressData"]
    start_row = 16
    item_num = 1
    total_due = 0
    border_style = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )
    alignment_center = Alignment(horizontal="center", vertical="center")
    # Insert today's date
    today = datetime.today().strftime("%Y-%m-%d")
    sheet["Q3"] = today

    # Fill in address
    sheet[f"B6"] = address_data["billToAddressLine1"]
    sheet[f"B7"] = address_data["billToAddressLine2"]
    sheet[f"B8"] = address_data["billToAddressLine3"]
    sheet[f"B9"] = address_data["billToAddressLine4"]
    sheet[f"N6"] = address_data["shipToAddressLine1"]
    sheet[f"N7"] = address_data["shipToAddressLine2"]
    sheet[f"N8"] = address_data["shipToAddressLine3"]
    sheet[f"N9"] = address_data["shipToAddressLine4"]

    for order_line in order_lines:
        # Insert a new row at the current start_row position
        sheet.insert_rows(start_row)

        # Add borders
        for cell in sheet[start_row]:
            cell.border = border_style

        # Merge cells
        sheet.merge_cells(
            start_row=start_row, start_column=2, end_row=start_row, end_column=3
        )
        sheet.merge_cells(
            start_row=start_row, start_column=5, end_row=start_row, end_column=6
        )
        sheet.merge_cells(
            start_row=start_row, start_column=7, end_row=start_row, end_column=8
        )

        # Center the content
        for cell in sheet[start_row]:
            cell.alignment = alignment_center

        # fill in the titles
        sheet[f"D12"] = order_line["ship_via"]
        sheet[f"I12"] = order_line["pay_terms"]
        
        # Populate the row with the order_line data
        balance = 0
        if order_line["balance"]:
            balance = int(order_line["balance"])
        
        surcharge_rate = 0
        if "surcharge_rate" in order_line and order_line["surcharge_rate"]:
            surcharge_rate = float(order_line["surcharge_rate"])
        
        discount_rate = 0
        if "discount" in order_line and order_line['discount']:
            discount_rate = float(order_line["discount"])

        basic_unit_price = 0
        if order_line["price"]:
            basic_unit_price = float(order_line["price"])
            if discount_rate:
                basic_unit_price *= (1-discount_rate*0.01)
        basic_unit_price = round(basic_unit_price, 2)
        sheet[f"M{start_row}"] = basic_unit_price

        unit_surcharge = basic_unit_price * 0.01 * surcharge_rate
        unit_surcharge = round(unit_surcharge, 2)

        unit_price = 0
        description = order_line["description"]
        if "include_surcharge" in order_line:
            unit_price = basic_unit_price + unit_surcharge
        else:
            if 'surcharge_line' in order_line:
                unit_price = unit_surcharge
                description = 'Surcharge ' + order_line['surcharge']
            else:
                unit_price = basic_unit_price
        
        total_quantity = int(order_line["quantity"]) + balance

        sheet[f"A{start_row}"] = item_num
        item_num += 1
        sheet[f"B{start_row}"] = order_line["customer_po"]
        sheet[f"D{start_row}"] = order_line["line_number"]
        sheet[f"E{start_row}"] = order_line["part_number"]
        sheet[f"I{start_row}"] = total_quantity
        sheet[f"J{start_row}"] = order_line["unit"]
        sheet[f"L{start_row}"] = order_line["material"]
        sheet[f"G{start_row}"] = description
        sheet[f"N{start_row}"] = str(surcharge_rate) + "%"
        sheet[f"O{start_row}"] = order_line["weight"]
        sheet[f"P{start_row}"] = unit_surcharge
        sheet[f"Q{start_row}"] = unit_price
        sheet[f"R{start_row}"] = order_line["total_price"]
        total_due += order_line["total_price"]

        # Move to the next row for the next order_line
        start_row += 1
    sheet[f"R{start_row}"] = total_due
    # Adjust the styles
    sheet.merge_cells(
        start_row=start_row + 1, start_column=1, end_row=start_row + 1, end_column=10
    )
    sheet[f"A{start_row+1}"].alignment = alignment_center
    # Save the workbook
    wb.save(destination_path)
